import openpyxl

wb = openpyxl.Workbook()
wb.create_sheet('Sheet1')
sheet = wb['Sheet1']
sheet['A1'] = 'Srilakshmi'
sheet['A2'] = 'Pavan'
sheet['A3'] = 'Bhuvan'
sheet['A4'] = 'Kasyap'

wb.create_sheet(index=1,title='My sheet')
wb.save('writing1.xlsx')
